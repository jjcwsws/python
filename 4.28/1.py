##import openpyxl as vb
##路径 = r'C:\Users\86138\OneDrive\桌面\生信\颜一决.xlsx'
##路径1 = r'C:\Users\86138\OneDrive\桌面\生信\颜一决1.xlsx'
##工作簿 = vb.load_workbook(路径)
##工作簿.remove(工作簿['刚刚复制的表1'])
##工作簿.save(路径)



##numdict = {1:"一",2:"二",3:"三",4:"四",5:"五",6:"六",7:"七",8:"八",9:"九",0:"零"} #个位数的字典
##digitdict = {1:"十",2:"百",3:"千",4:"万"} #位称的字典
## 
## 
##def maxdigit(number,count):
##    num = number//10 #整除是//
##    if num != 0:
##        return maxdigit(num,count+1) #加上return才能进行递归
##    else:
##        digit_num = number%10 #digit_num是最高位上的数字
##        return count,digit_num #count记录最高位
## 
##def No2Cn(number):
##    max_digit,digit_num = maxdigit(number,0)
## 
##    temp = number
##    num_list = [] #储存各位数字（最高位的数字也可以通过num_list[-1]得到
##    while temp > 0:
##        position = temp%10
##        temp //= 10 #整除是//
##        num_list.append(position)
## 
##    chinese = ""
##    if max_digit == 0: #个位数
##        chinese = numdict[number]
##    elif max_digit == 1: #十位数
##        if digit_num == 1: #若十位上是1，则称为“十几”，而一般不称为“一十几”（与超过2位的数分开讨论的原因）
##            chinese = "十"+numdict[num_list[0]]
##        else:
##            chinese = numdict[num_list[-1]]+"十"+numdict[num_list[0]]
##    elif max_digit > 1: #超过2位的数
##        while max_digit > 0:
##            if num_list[-1] != 0: #若当前位上数字不为0，则加上位称
##                chinese += numdict[num_list[-1]]+digitdict[max_digit]
##                max_digit -= 1
##                num_list.pop(-1)
##            else: #若当前位上数字为0，则不加上位称
##                chinese += numdict[num_list[-1]]
##                max_digit -= 1
##                num_list.pop(-1)
##        chinese += numdict[num_list[-1]]
##        
##    while chinese.endswith("零") and len(chinese) > 1: #个位数如果为0，不读出
##        chinese = chinese[:-1]
##    if chinese.count("零") > 1: #中文数字中最多只有1个零
##        count_0 = chinese.count("零")
##        chinese = chinese.replace("零","",count_0-1)
##    return chinese

#新建一个七月1日到七月31日的工作表（sheet）
import openpyxl as vb
路径 = r'C:\Users\86138\OneDrive\桌面\生信\第一次测试题.xlsx'
工作簿 = vb.Workbook(路径)
for i in range(1,32):
    工作簿.create_sheet('七月' + str(i) +'日' )
工作簿.save(路径)
