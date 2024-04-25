#import openpyxl as vb
#路径 = r'C:\Users\86138\OneDrive\桌面\生信\颜一决.xlsx'
#工作簿 = vb.load_workbook(路径)
#显示所有工作表 = 工作簿.worksheets
#print(显示所有工作表)

#import openpyxl as vb
#路径 = r'C:\Users\86138\OneDrive\桌面\生信\颜一决.xlsx'
#工作簿 = vb.load_workbook(路径)
#显示所有工作表 = 工作簿.worksheets
#for i in 显示所有工作表:
#   print(i.title)
#
#第一步，导入模块
#第二部，指定路径
#第三步，打开路径


#第四步，首先指定该工作表是什么
import openpyxl as vb
路径 = r'C:\Users\86138\OneDrive\桌面\生信\颜一决.xlsx'
路径1 = r'C:\Users\86138\OneDrive\桌面\生信\颜一决.xlsx'
工作簿 = vb.load_workbook(路径)
工作表 = 工作簿['三月']
工作簿.remove(工作表)
工作簿.save(路径1)
