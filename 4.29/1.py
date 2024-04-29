##import openpyxl as vb
##路径 = r'C:\Users\86138\OneDrive\桌面\生信\第一次测试题.xlsx'
##工作簿 = vb.Workbook(路径)
##for i in range(1,32):
####    工作簿.create_sheet(f"7月{i}日")   此处不用考虑i是数值型还是字符串型
####    工作簿.create_sheet('七月' + str(i) +'日' )
##工作簿.save(路径)


##import openpyxl as vb
##路径 = r'C:\Users\86138\OneDrive\桌面\生信\第一次测试题.xlsx'
##工作簿 = vb.load_workbook(路径)
##显示所有工作表 = 工作簿.worksheets
##for i in 显示所有工作表:
##    i.title = '北京' + '-' +  i.title
##工作簿.save(路径)


import openpyxl as vb
路径 = r'C:\Users\86138\OneDrive\桌面\生信\第一次测试题.xlsx'
工作簿 = vb.load_workbook(路径)
显示所有工作表 = 工作簿.worksheets
for i in 显示所有工作表:
    if i.title.split()
